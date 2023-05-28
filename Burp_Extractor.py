# Importing necessary modules
import re
import xml.etree.ElementTree as ET
import base64
import json
import os
import multiprocessing
import time
import argparse
import traceback
import urllib.parse
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning
import math
from collections import Counter
import warnings
import spacy

# Defining some colors for output formatting
GREEN = "\033[32m"
RESET = "\033[0m"
RED = "\033[31m"
BLUE = "\033[1;34m"
ORANGE = "\033[1;33m"
MAGENTA = "\033[1;35m"


def cleaning(host, lines):

    try:
        # Installation
        # pip install spacy
        # python -m spacy download en_core_web_sm
        nlp = spacy.load("en_core_web_sm")

        regexes = [
            r".{100,}",  # Ignore lines with more than 100 characters (overly specific)
            r"[0-9]{4,}",  # Ignore lines with 4 or more consecutive digits (likely an id)
            r"[0-9]{3,}$",  # Ignore lines where the last 3 or more characters are digits (likely an id)
            r"[a-z0-9]{32}",  # Likely MD5 hash or similar
            r"\b[A-Z]{2,}\w*\b",  # Matches uppercase strings with two or more characters
            r"\b[A-Z0-9]{5,}\b",  # Matches strings with five or more uppercase letters or digits
            r"\b[A-Z]{2,}\w*\b" # Matches a word beginning with 2 or more Uppercase letters 
            r"\b[BCD]{3,}\b",  # Matches strings with three or more consecutive B, C, or D characters
            r"[0-9]+[A-Z0-9]{5,}",  # Number followed by 5 or more numbers and uppercase letters (almost all noise)
            r"\/.*\/.*\/.*\/.*\/.*\/.*\/",  # Ignore lines more than 6 directories deep (overly specific)
            r"\w{8}-\w{4}-\w{4}-\w{4}-\w{12}",  # Ignore UUIDs
            r"[0-9]+[a-zA-Z]+[0-9]+[a-zA-Z]+[0-9]+",  # Ignore multiple numbers and letters mixed together (likely noise)
            r"\.(png|jpg|jpeg|gif|svg|bmp|ttf|avif|wav|mp4|aac|ajax|css|all)$",  # Ignore low-value file types
            r"^$",  # Ignores blank lines
            r"[^a-zA-Z0-9\s_.-]+",  # Remove non-alphanumeric characters except underscore, dash, and dot at the beginning of a line
        ]

        print(f'{BLUE}\n[+] Cleaning Wordlist, please wait this make take a while..{RESET}')
        print(f'{BLUE}\n[+] Crazy A** calculations is happening right now man, chill..{RESET}')

        original_size = len(lines)

        # Apply regexes to remove lines
        for regex in regexes:
            pattern = re.compile(regex)
            lines = [line for line in lines if not pattern.search(line)]

        # Remove lines starting with digits
        lines = [line for line in lines if not re.search(r"^[0-9]", line)]

        # Remove lines that contain only a single character
        lines = [line for line in lines if len(line.strip()) > 1]

        # Sort and remove duplicates
        lines = sorted(set(lines))

        second_lines = [line.replace(".js.map", "").replace(".js", "").replace(".map", "").replace(".min.js", "").replace(".min.map", "").replace(".min", "")
                for line in lines
                if ('-' in line or '.' in line or '_' in line or line.endswith((".js.map", ".js", ".map", ".min.js", ".min.map")))
                and not (line.endswith(".js.map") or line.endswith(".js") or line.endswith(".map") or line.endswith(".min.js") or line.endswith(".min.map")) or line.endswith(".min")]
        
        lines = [line for line in lines if any(token.is_alpha and not token.is_stop and len(token.text) > 1 for token in nlp(line.lower()))]

        # Calculate changes
        new_size = len(lines) + len(second_lines)
        removed = original_size - new_size

        print(f"{BLUE}\n[+] Removed {removed} lines{RESET}")
        print(f"{BLUE}\n[+] Wordlist is now {new_size} lines{RESET}")
        print(f"{BLUE}\n[+] Done{RESET}")

        with open(f'{host}\{host}_wordlist.txt', 'w', encoding="utf-8") as f:
            
            for item in lines:
                f.write(f"{item}\n")

            for second_item in second_lines:
                f.write(f"{second_item}\n")

        print(f'{GREEN}\n[+] Wordlist saved to {host}\{host}_wordlist.txt{RESET}')  

    except:
        print(f"{RED}\n[+] Please install Spacy by issuing these commands in the command line:\n[+] pip install spacy\n[+] python -m spacy download en_core_web_sm{RESET}")

def entropy(string):
    #"Calculates the Shannon entropy of a string"
    # get probability of chars in string
    prob = [float(string.count(c)) / len(string) for c in dict.fromkeys(list(string))]

    # calculate the entropy
    entropy = - sum([p * math.log(p) / math.log(2.0) for p in prob])

    return entropy

def wordlist_creator(file, host):

    tree = ET.parse(file)
    root = tree.getroot()
    wordlist = []

    print(f"{BLUE}\n[+] Please wait, it might take a few minutes...{RESET}")

    for i in root:

        # preserve subdomains, file/dir names with . - _
        wordlist += re.split('\/|\?|&|=', i[1].text)

        # get subdomain names and break up file names
        wordlist += re.split('\/|\?|&|=|_|-|\.|\+', i[1].text)

        # get words from cookies, headers, POST body requests
        wordlist += re.split('\/|\?|&|=|_|-|\.|\+|\:| |\n|\r|"|\'|<|>|{|}|\[|\]|`|~|\!|@|#|\$|;|,|\(|\)|\*|\|', urllib.parse.unquote(base64.b64decode(i[8].text)))

        # response
        if i[12].text is not None:
            wordlist += re.split('\/|\?|&|=|_|-|\.|\+|\:| |\n|\r|\t|"|\'|<|>|{|}|\[|\]|`|~|\!|@|#|\$|;|,|\(|\)|\*|\^|\\\\|\|', urllib.parse.unquote(base64.b64decode(i[12].text)))

    auxiliaryList = list(set(wordlist))
    final = []

    for word in auxiliaryList:
        if word.isalnum() or '-' in word or '.' in word or '_' in word:
            en = entropy(word)
            # remove "random strings" that are high entropy
            if en < 4.4:
                final.append(word)

    final.sort()

    print(f"{BLUE}\n[+] Wordlist is {len(final)} lines{RESET}")

    cleaning(host, final)

def avgEntropyByChar(en, length):
    # calulate "average" entropy level
    return en / length    

def is_any_process_alive(processes):
    return True in [p.is_alive() for p in processes]

def adjust_column_widths_disk(filepath):
    workbook = load_workbook(filepath)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.3
            sheet.column_dimensions[column].width = adjusted_width

    workbook.save(filepath)

def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.3
        sheet.column_dimensions[column].width = adjusted_width
        
def postMan(file):

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()

    headers_list = []
    unique_path = set()
    unique_name_path = set()
    counter = 0

    for i in root:
        host = i.find('host').text
        break

    domain_output = i.find('host').text

    POST_Requests = []        

    postman = {

        "info": {
            "_postman_id": "my-postman-id",
            "name": f"{host} API Endpoints",
            "description": "API Endpoints Documentation - Generated by Clear Gate Cyber Security",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "items": []
        }

    # Looping through each request/response
    for i in root:
        # Searching for responses only
        response = i.find('response').text
        if response is None:
            continue
        # Decoding the response
        content = base64.b64decode(response)
        # Filtring the Content-Type
        content_type_pattern = re.compile(b'Content-Type: .*?(?:\r\n|\r|\n)', re.DOTALL)
        match = content_type_pattern.search(content)

        if match:
            content_type_header = match.group().decode('latin-1')

        method = i.find('method').text
        
        # Searching for responses only
        request = i.find('request').text
        # Decoding the response
        content_request = base64.b64decode(request)
        match_request = content_type_pattern.search(content_request)
        content_type_header_request = None

        if match_request is not None:
            content_type_header_request = match_request.group(0).decode('latin-1')
        if method == 'POST':

            if match:
                content_type_header = match.group().decode('utf-8')
                # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                if content_type_header is not None and 'application/json' in content_type_header or 'application/xml' in content_type_header:
                   
                    path = i.find('path').text
                    name_path = re.sub(r'\?.*', '', path)

                    if name_path in unique_name_path:
                        continue
                    else:
                        unique_name_path.add(name_path)

                    if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                        continue
                    
                    else:
                        path = path
                        if path.endswith("/"):
                            path = path[:-1]

                        domain = i.find('host').text
                        host = i.find('host').text
                        protocol = i.find('protocol').text
                        
                        # Searching for requests only
                        request = i.find('request').text
                        # Decoding the request
                        content_request = base64.b64decode(request)
                        content_request = content_request.decode('latin-1')
                        headers_list = re.findall(r'(?P<name>.*?): (?P<value>.*?)\r\n', content_request)
                        headers_list = [{'key': key, 'value': value} for key, value in headers_list if value]

                        url = i.find('url').text
                        # Searching for requests only
                        request = i.find('request').text
                        # Decoding the request
                        content_request = base64.b64decode(request)
        
                        body = content_request.split(b'\r\n\r\n', 1)[1].decode('latin-1')
                        # create a dictionary to represent the POSTMAN collection
                        loop_collection = {
                            "name": name_path,
                            "request": {
                                "method": "POST",
                                "header": headers_list,
                                "url": {
                                    "raw": url,
                                    "host": [
                                        f"{protocol}://{host}"
                                    ],
                                    "path": [
                                        path
                                    ]
                                },
                                "description": "Not documented",
                                "body": {
                                    "mode": "raw",
                                    "raw": body
                                }
                            },
                            "response": []
                        }
                        
                        
                        POST_Requests.append(loop_collection)
                        

                        if path not in unique_path and body != '[]' and content_type_header_request is not None:
                            unique_path.add(path)

                            # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                            if 'application/json' in content_type_header_request:
                                mode = "raw"
                                kv_pairs = body
                                
                            elif 'application/xml' in content_type_header_request:
                                mode = "raw"
                                kv_pairs = body

                            elif 'application/x-www-form-urlencoded' in content_type_header_request:
                                mode = "urlencoded"
                                # URL decode the string
                                body = urllib.parse.unquote(body)
                                
                                # Split the string by "&" to get a list of key-value pairs
                                pairs_list = body.split("&")

                                # Initialize an empty list to store the key-value pairs
                                kv_pairs = []

                                # Loop through each pair and split it by "=" to get the key and value separately
                                for pair in pairs_list:
                                    if "=" in pair:
                                        values = pair.split('=', 1)
                                        if values[1] == "":
                                            
                                            values.insert(1, f"<{values[0]}>")
                                        # Append the key-value pair to the kv_pairs list
                                        kv_pairs.append({"key": values[0], "value": values[1]})

                            else:
                                mode = "raw"
                                kv_pairs = body

                            # create a dictionary to represent the POSTMAN collection
                            loop_collection = {
                                "name": name_path,
                                "request": {
                                    "method": "POST",
                                    "header": headers_list,
                                    "url": {
                                        "raw": url,
                                        "host": [
                                            f"{protocol}://{host}"
                                        ],
                                        "path": [
                                            path
                                        ]
                                    },
                                    "description": "Not documented",
                                    "body": {
                                        "mode": mode,
                                        mode: kv_pairs
                                    }
                                },
                                "response": []
                            }

                            if str(loop_collection['request']['url']['path']) in str(POST_Requests):
                                continue

            mime_types = ['application/x-www-form-urlencoded', 'multipart/form-data', 'application/json', 'application/xml', 'text/plain']
            
            if match_request:
                
                for mime in mime_types:
                    if mime in content_type_header_request:

                        path = i.find('path').text

                        name_path = re.sub(r'\?.*', '', path)
                                            
                        if name_path in unique_name_path:
                            continue
                        else:
                            unique_name_path.add(name_path)

                        if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                            continue
                        
                        else:
                            path = path
                            if path.endswith("/"):
                                path = path[:-1]

                            host = i.find('host').text
                            protocol = i.find('protocol').text
                            url = i.find('url').text

                            # Searching for requests only
                            request = i.find('request').text
                            # Decoding the request
                            content_request = base64.b64decode(request)
                            content_request = content_request.decode('latin-1')
                            headers_list = re.findall(r'(?P<name>.*?): (?P<value>.*?)\r\n', content_request)
                            headers_list = [{'key': key, 'value': value} for key, value in headers_list if value]

                            # Searching for requests only
                            request = i.find('request').text
                            # Decoding the request
                            content_request = base64.b64decode(request)
            
                            body = content_request.split(b'\r\n\r\n', 1)[1].decode('latin-1')
                            
                            if path not in unique_path and body != '[]':
                                unique_path.add(path)

                                # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                                if 'application/json' in content_type_header_request:
                                    mode = "raw"
                                    kv_pairs = body
                                    
                                elif 'application/xml' in content_type_header_request:
                                    mode = "raw"
                                    kv_pairs = body

                                elif 'application/x-www-form-urlencoded' in content_type_header_request:
                                    mode = "urlencoded"
                                    # URL decode the string
                                    body = urllib.parse.unquote(body)
                                    
                                    # Split the string by "&" to get a list of key-value pairs
                                    pairs_list = body.split("&")

                                    # Initialize an empty list to store the key-value pairs
                                    kv_pairs = []

                                    # Loop through each pair and split it by "=" to get the key and value separately
                                    for pair in pairs_list:
                                        if "=" in pair:
                                            values = pair.split('=', 1)
                                            if values[1] == "":
                                                
                                                values.insert(1, f"<{values[0]}>")
                                            # Append the key-value pair to the kv_pairs list
                                            kv_pairs.append({"key": values[0], "value": values[1]})

                                else:
                                    mode = "raw"
                                    kv_pairs = body

                                # create a dictionary to represent the POSTMAN collection
                                loop_collection = {
                                    "name": name_path,
                                    "request": {
                                        "method": "POST",
                                        "header": headers_list,
                                        "url": {
                                            "raw": url,
                                            "host": [
                                                f"{protocol}://{host}"
                                            ],
                                            "path": [
                                                path
                                            ]
                                        },
                                        "description": "Not documented",
                                        "body": {
                                            "mode": mode,
                                            mode: kv_pairs
                                        }
                                    },
                                    "response": []
                                }
                                
                                POST_Requests.append(loop_collection)

        if method == 'GET':

            if match:
                content_type_header = match.group().decode('utf-8')
                # if Content-Type is equals to JSON/XML it will be added to the xlsx file
                if 'application/json' in content_type_header or 'application/xml' in content_type_header:

                    path = i.find('path').text
                    if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                        continue
                    
                    else:

                        path = i.find('path').text
                        second_path = path = i.find('path').text
                        name_path = re.sub(r'\?.*', '', path)

                        if name_path in unique_name_path:
                            continue
                        else:
                            unique_name_path.add(name_path)
                        
                        if "?" in path:
                            path = path.split("?")
                            query_parameters = path[1]
                            path = path[0]
                        if "?" not in path:
                            path = path
                            
                        domain_output = i.find('host').text
                        host = i.find('host').text
                        protocol = i.find('protocol').text
                        url = i.find('url').text
                        
                        if path in unique_path:
                            continue
                        else:
                            unique_path.add(path)
                            counter += 1

                        # create a dictionary to represent the POSTMAN collection
                        loop_collection = {
                            "name": f"{name_path}",
                            "request": {
                                "method": "GET",
                                "header": headers_list,
                                "url": {
                                    "raw": url,
                                    "host": [
                                        f"{protocol}://{host}"
                                    ],
                                    "path": [
                                        path
                                    ]
                                },

                                "query": [],
                            },
                            "response": []
                        }
                        
                        if "?" in second_path:

                            # URL decode the string
                            parameters = urllib.parse.unquote(query_parameters)
                            # Split the string by "&" to get a list of key-value pairs
                            pairs_list = parameters.split("&")

                            # Initialize an empty list to store the key-value pairs
                            kv_pairs = []

                            # Loop through each pair and split it by "=" to get the key and value separately
                            for pair in pairs_list:
                                if "=" in pair:
                                    values = pair.split('=', 1)
                                    if values[1] == "":
                                        values.insert(1, f"<{values[0]}>")

                                    #Append the key-value pair to the kv_pairs list
                                    kv_pairs.append({"key": values[0], "value": values[1]})

                            loop_collection["request"]["url"]["query"] = kv_pairs

                        # Searching for requests only
                        request = i.find('request').text
                        # Decoding the request
                        content_request = base64.b64decode(request)
                        content_request = content_request.decode('latin-1')
                        headers_list = re.findall(r'(?P<name>.*?): (?P<value>.*?)\r\n', content_request)
                        headers_list = [{'key': key, 'value': value} for key, value in headers_list if value]

                        postman["items"].append(loop_collection)

    for post in POST_Requests:
        counter += 1
        postman["items"].append(post)

    postman['info'].update({"description": f"Total {counter} API Endpoints - Generated by Clear Gate Cyber Security"})

    if counter == 0:
        print(f'\n{RED}[-] No API Endpoints found in order to generate JSON file for Postman.{RESET}') 

    if counter > 0:
        # convert the dictionary to a JSON string and print it
        json_collection = json.dumps(postman)

        if not os.path.exists(host):
            os.system(f"mkdir {host}")

        with open(f"{host}\{host}.json", "w") as f:
            f.write(f"{json_collection}")
            print(f'{GREEN}\n[+] {domain_output}.json was created in your current directory!{RESET}')
            print(f'{GREEN}[+] You can open it with{RESET} {ORANGE}Postman!{RESET}')    

# Function for extracting API Endpoints from Burp Response based on XML/JSON Content-Type (Soap/REST)
def js_file(file, wb, sheet):

    js_list = []

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()

    for i in root:

        domain = i.find('host').text

        url = i.find('url').text

        if url.endswith('.js') or 'js?' in url or url.endswith('.map') or 'map?' in url:
            print(f'\n{MAGENTA}[+] Testing {url}{RESET}')
            if 'js?' in url or 'map?' in url:
                if url.endswith('/') or url.endswith('\\'):
                    url = url[:-1]
                url = url.split('?')[0]    

            js_list.append(url)

    data = []

    new_list = []

    for js_file in js_list:

        if js_file not in new_list:
            new_list.append(js_file)

        elif js_file in new_list:
            continue

    for js in new_list:
        data.append([js])

    for row in data:
        sheet.append(row)
        # Setting font of the cell to Calibri 14
        row = sheet.max_row
        sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)

    if not os.path.exists(domain):
        os.system(f"mkdir {domain}")

    # Adjust column widths
    adjust_column_widths(sheet)
    wb.save(f'{domain}\{domain}_JS_Files.xlsx')    

    print(f'{GREEN}\n[+] {domain}_JS_Files.xlsx was created in your current directory!{RESET}')

# Function for extracting API Endpoints from Burp Response based on XML/JSON Content-Type (Soap/REST)
def json_file(file, wb):

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()
    
    data = []

    # Looping through each request/response
    for i in root:
        # Searching for responses only
        response = i.find('response').text
        url = i.find('url').text
        if response is None:
            continue
        # Decoding the response
        content = base64.b64decode(response)
        # Filtring the Content-Type
        content_type_pattern = re.compile(b'Content-Type: .*?(?:\r\n|\r|\n)', re.DOTALL)
        match = content_type_pattern.search(content)

        if match:
            content_type_header = match.group().decode('utf-8')
            # if Content-Type is equals to JSON/XML it will be added to the xlsx file
            if 'application/json' in content_type_header:
                
                path = i.find('path').text
                
                if path.endswith(".json") or path.endswith(".json?"):
                
                    #path = path.split("?")[0]
                    domain = i.find('host').text
                    data.append([domain, path, url])
                else:
                    continue

    if len(data) > 0:
        # Sorting and removing duplicates
        data = sorted(list(set([tuple(row) for row in data])))
    
        if args.directory:
        
            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break
            
            sheet, wb = create_worksheet_json(host, "JSON Files")

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            sheet, wb = create_worksheet_json(host, "JSON Files")

        for row in data:
            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            # Removing duplicates and sorting rows

            if not os.path.exists(domain):
                os.system(f"mkdir {domain}")

            # Adjust column widths
            adjust_column_widths(sheet)
            wb.save(f'{domain}\{domain}_JSON_Files.xlsx')    
        
        print(f'{GREEN}\n[+] {domain}_JSON_Files.xlsx was created in your current directory!{RESET}')

    else:
        print(f'\n{RED}[-] No JSON Files found.{RESET}')    
 
# Function for extracting API Endpoints from Burp Response based on XML/JSON Content-Type (Soap/REST)
def bitrix(file):

    # Declaring an XML object
    tree = ET.parse(file)
    root = tree.getroot()
    data = []

    # Looping through each request/response
    for i in root:
        # Searching for responses only
        response = i.find('response').text
        if response is None:
            continue
        # Decoding the response
        content = base64.b64decode(response)
        # Filtring the Content-Type
        content_type_pattern = re.compile(b'Content-Type: .*?(?:\r\n|\r|\n)', re.DOTALL)
        match = content_type_pattern.search(content)

        # Searching for responses only
        request = i.find('request').text
        if request is None:
            continue

        # Decoding the response
        content_request = base64.b64decode(request)

        match_request = content_type_pattern.search(content_request)

        if match_request is not None:
            content_type_header_request = match_request.group(0).decode('latin-1')
  
        if match:
            content_type_header = match.group().decode('utf-8')
            # if Content-Type is equals to JSON/XML it will be added to the xlsx file
            if 'application/json' in content_type_header or 'application/xml' in content_type_header:
                
                path = i.find('path').text

                if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                    continue
                
                path = path.split("?")[0]
            
                method = i.find('method').text
                domain = i.find('host').text
                data.append([domain, path, method])

        mime_types = ['application/x-www-form-urlencoded', 'multipart/form-data', 'application/json', 'application/xml', 'text/plain']
        
        if match_request:
            
            for mime in mime_types:
                if mime in content_type_header_request:
                    
                    path = i.find('path').text
                    if path.endswith("map") or path.endswith("map?") or path.endswith("json") or path.endswith("json?"):
                        continue

                    path = path.split("?")[0]
                    method = i.find('method').text
                    domain = i.find('host').text
                    data.append([domain, path, method])

    if len(data) > 0:
        
        # Sorting and removing duplicates
        data = sorted(list(set([tuple(row) for row in data])))

        if args.directory:

            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break
            
            sheet, wb = create_worksheet_bitrix(host, "API_Endpoints_Bitrix")

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(file)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            sheet, wb = create_worksheet_bitrix(host, "API_Endpoints_Bitrix")

        # Sorting and removing duplicates
        for row in data:

            sheet.append(row)
            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            # Removing duplicates and sorting rows

        if not os.path.exists(domain):
            os.system(f"mkdir {domain}")     
        
        # Adjust column widths
        adjust_column_widths(sheet)
        wb.save(f'{domain}\{domain}_API_Endpoints_Bitrix.xlsx')
        print(f'{GREEN}\n[+] {domain}_API_Endpoints.xlsx was created in your current directory!{RESET}')


    else:
        print(f'\n{RED}[-] No API Endpoints found (XML/JSON Content-Type) with Bitrix method.{RESET}')    


def parse_args():
    banner = """ __        __   __      ___     ___  __        __  ___  __   __
|__) |  | |__) |__)    |__  \_/  |  |__)  /\  /  `  |  /  \ |__)   
|__) \__/ |  \ |       |___ / \  |  |  \ /~~\ \__,  |  \__/ |  \ 

Developed by Sagiv
Clear Gate - Cyber Security                                                                                                                           """ 
    parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter, description=banner)
    parser.add_argument('-f', '--file', type=str, required=False, help='Burp File (Right Click on the domain in the Target Scope and select save selected items and select Base64 encode).')
    parser.add_argument('-dr', '--directory', type=str, required=False, help='Directroy containing all Burp Suite output files.')
    parser.add_argument('-a', '--all', required=False, action="store_true", help='Use all methods below - Can be be slow depends on the size of the project, so leave it running in the background.')
    parser.add_argument('-b', '--bitrix', required=False, action="store_true", help='Generate API Endpoints to Excel file based on JSON/XML Content-Type via Burp Response - Fast and *Recommended* for Bitrix24 Task.')
    parser.add_argument('-p', '--postman', required=False, action="store_true", help='Collect --bitrix result to JSON file with body and parameters for Postman Application - Fast and *Recommended*.')
    parser.add_argument('-w', '--wordlist', required=False, action="store_true", help="Create a tailored wordlist for your target (Based on Request/Responses including Headers/Cookies and body - Can be up to 2 minutes, but very *recommended*.")
    parser.add_argument('-d', '--domain', required=False, action="store_true", help='Collect Subdomains based on Burp response via REGEX to Excel file - Fast.')
    parser.add_argument('-j', '--json', required=False, action="store_true", help='Collect JSON files based on Burp response via REGEX to Excel file - Fast.')
    parser.add_argument('-J', '--js', required=False, action="store_true", help='Collect JS/MAP URLs based on Burp response via REGEX to Excel file - Fast.')
    parser.add_argument('-i', '--api', required=False, action="store_true", help='Collect APIs and PATHs based on Burp response via REGEX to Excel file - Might be slow depends on the size of the project.')
    parser.add_argument('-s', '--secrets', required=False, action="store_true", help="Collect Secrets (AWS/Google keys, etc' - A lot of False-Positive) based on Burp response via REGEX to Excel file - Might be slow depends on the size of the project.")
    parser.add_argument('-u', '--urls', required=False, action="store_true", help='Collect URLs based on Burp response via REGEX to Excel file - Might be slow depends on the size of the project.')
    parser.add_argument('-t', '--threads', required=False, type=int, default=os.cpu_count(), help='Number of threads run in parallel (Default is the number of your CPU Cores).')
    parser.add_argument('-v', '--verbose', required=False, action="store_true", help='If set, output will be printed to the screen with colors.')
    return parser.parse_args()

def create_worksheet(host, string):
    
    # Creating a new Workbook object
    wb = Workbook()
    
    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = f"{host}_{string}"

    sheet['A1'] = 'URL Tested'
    sheet['B1'] = 'Regex'
    sheet['C1'] = 'Matched Pattern'
    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font

    return sheet, wb

def create_worksheet_main(string):
    
    # Creating a new Workbook object
    wb = Workbook()
    
    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = string

    sheet['A1'] = 'URL Tested'
    sheet['B1'] = 'Regex'
    sheet['C1'] = 'Matched Pattern'
    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font

    return sheet, wb

def create_worksheet_bitrix(host, string):
        
        # Creating a new Workbook object
        wb = Workbook()
        
        # Creating a sheet for the matched patterns and setting the font of the header row
        sheet = wb.active
        sheet.title = f"{host}_{string}"

        # Writing data to sheet
        sheet = wb.active
        sheet.title = "API Endpoints"
        sheet['A1'] = 'HOST'
        sheet['B1'] = 'ENDPOINT'
        sheet['C1'] = 'METHOD'
        sheet['D1'] = 'TESTED?'
        header_font = Font(name='Calibri', size=20, bold=True)
        sheet['A1'].font = header_font
        sheet['B1'].font = header_font
        sheet['C1'].font = header_font
        sheet['D1'].font = header_font

        return sheet, wb

def create_worksheet_json(host, string):

    # Creating a new Workbook object
    wb = Workbook()
    
    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = f"{host}_{string}"

    # Writing data to sheet
    sheet = wb.active
    sheet.title = "JSON Files"
    sheet['A1'] = 'HOST'
    sheet['B1'] = 'JSON File'
    sheet['C1'] = 'URL'

    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font
    sheet['B1'].font = header_font
    sheet['C1'].font = header_font

    return sheet , wb

def create_worksheet_js(string):

    # Creating a new Workbook object
    wb = Workbook()
    
    # Creating a sheet for the matched patterns and setting the font of the header row
    sheet = wb.active
    sheet.title = string

    # Writing data to sheet
    sheet = wb.active
    sheet.title = "JS Files"
    sheet['A1'] = 'JS/MAP File'

    header_font = Font(name='Calibri', size=20, bold=True)
    sheet['A1'].font = header_font

    return sheet , wb

def match(regex, content, url, host, sheet, wb, matched_patterns, string, static_files, args, final_xlsx):
    
    data = []  

    try:
        # Iterating over the given regex patterns to search for matches
        for key, value in regex.items():
            if args.verbose:
                # Printing the URL being tested
                print(f'Testing Regex: {RED}{key}: {value}{RESET}')
                print(f'Testing URL: {RED}{url}{RESET}')

            # Searching for a match of the current regex pattern
            pattern = re.compile(value, re.IGNORECASE | re.MULTILINE)

            match = re.findall(pattern, content)
            

            if match:
                flag = False
                for matched_pattern in match:

                    # Only add unique matched patterns to the set
                    if matched_pattern not in matched_patterns:
                        if string == 'Path_and_Endpoints':
                            parts = matched_pattern.split("/")
                            if "." in parts[-1]:
                                continue
                           
                            threshold = 4
                            entropy_string = str(matched_pattern)[1:]  # exclude the first forward slash
                            # Calculate the entropy of the string
                            prob = [float(entropy_string.count(c)) / len(entropy_string) for c in dict.fromkeys(list(entropy_string))]
                            # calculate the entropy
                            entropy = - sum([p * math.log(p) / math.log(2.0) for p in prob])
                           
                            try:
                                flag = False
                                if entropy < threshold:
                                    
                                    if "//" in str(matched_pattern):
                                         continue

                                    if "\/\/" in str(matched_pattern):
                                        continue
                                    
                                    if matched_pattern.count('/') < 2:
                                        continue

                                    if '\/\/' in str(matched_pattern):
                                        continue
                                    
                                    if "//" in str(matched_pattern):
                                        continue
                                    
                                    for staticFile in static_files:
                                        if staticFile.lower() in str(matched_pattern.lower()):
                                            flag = True
                                            continue
                                    
                                    if flag:
                                        continue

                                    split_string = str(matched_pattern).split("/")
                                    
                                    for split in split_string:
                                        
                                        if "-" in split or "_" in split:
                                            continue
                                        
                                        if len(split) == 1:
                                            flag = True
                                            continue

                                    if flag:
                                        continue
                                    
                                if entropy > threshold:
                                    continue

                            except Exception as error:
                                print(error) 
                                traceback.print_exc()

                        matched_patterns.add(matched_pattern)
                        matched_pattern = list(matched_pattern)

                        if len(matched_pattern) != 0:
                            if len(matched_pattern) > 2:
                                if matched_pattern[2].endswith('\\'):
                                    matched_pattern[2] = matched_pattern[2][:-1]
                                if '\\/' in matched_pattern[2]:
                                    matched_pattern[2] = matched_pattern[2].replace('\\/', '/')
                            if 'http' == matched_pattern[0]:
                                matched_pattern[0] = 'http://'

                            elif 'https' == matched_pattern[0]:
                                matched_pattern[0] = 'https://'

                            elif 'ftp' == matched_pattern[0]:
                                matched_pattern[0] = 'ftp://'

                            matched_pattern = "".join(matched_pattern)
                            if matched_pattern.endswith("?"):
                                matched_pattern = matched_pattern[:-1]

                            if args.verbose:
                                print(f'Matched regex: {GREEN}{key}: {value}{RESET} with pattern: {GREEN}{matched_pattern}{RESET}')

                            if '\\/' in matched_pattern:
                                matched_pattern = matched_pattern.replace('\\/', '/')
                                
                            # Only append unique regex matches to data
                            if matched_pattern not in [row[2] for row in data]:
                                if host in url:
                                    data.append([url, f"{key}: ({value})", matched_pattern])

            elif not match:
                if args.verbose:
                    print(f'{RED}\n[-] No match has been found for this regex.{RESET}')

        # Removing duplicates and sorting rows
        data = sorted(list(set([tuple(row) for row in data])))  

        for row in sorted(data):
            sheet.append(row)

            # Setting font of the cell to Calibri 14
            row = sheet.max_row
            sheet.cell(row=row, column=1).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=2).font = Font(name='Calibri', size=14)
            sheet.cell(row=row, column=3).font = Font(name='Calibri', size=14)
            
            if not os.path.exists(host):
                os.system(f"mkdir {host}")

            final_xlsx.append(f"{host}\{host}_{string}.xlsx")
            wb.save(f"{host}\{host}_{string}.xlsx")
            wb.close()

        return host

    except Exception as error:
        # print the error message and traceback
        print(error)
        traceback.print_exc()
        exit(1)
    
def main(file, tool_method, sheet, wb, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx):

    # Create an XML Object
    tree = ET.parse(file)
    root = tree.getroot()
    endpoint_check = ""
    final_host = None
    empty_list = []
    empty_url_list = []
    flag = False

    # Checks if file imoprted from HTTP History to determine different hosts

    for i in root:

        response = i.find('response').text
        
        if response is None:
            continue

        content = base64.b64decode(response)
        content = content.decode('latin-1')
        hostOne = i.find('host').text
        break
    
    counter = 0
    for i in root:
        counter += 1
        response = i.find('response').text
        
        if response is None:
            continue

        content = base64.b64decode(response)
        content = content.decode('latin-1')
        hostTwo = i.find('host').text    

        if hostTwo != hostOne:
            flag = True
            break

        elif counter > 15:
            break    
    
    for i in root:

        response = i.find('response').text
        
        if response is None:
            continue

        content = base64.b64decode(response)
        content = content.decode('latin-1')
        unique_host = i.find('host').text

        if flag:

            if tool_method == "Secrets":
                sheet, wb = create_worksheet(unique_host, "Secrets")
            elif tool_method == "Path_and_Endpoints":
                sheet, wb = create_worksheet(unique_host,"Path_and_Endpoints")
            elif tool_method == "URLs":
                sheet, wb = create_worksheet(unique_host, "URLs")
            elif tool_method == "Sub-Domains":
                sheet, wb = create_worksheet(unique_host, "Sub-Domains")

        url = i.find('url').text

        if url not in empty_url_list:
            empty_url_list.append(url)
            
        elif url in empty_url_list:   
            continue

        unique_path = i.find('path').text

        if unique_path not in empty_list:
            empty_list.append(unique_path)
        
        elif unique_path in empty_list:   
            continue

        parsed_url = urlparse(url)
        domain_parts = parsed_url.netloc.split(".")
        if len(domain_parts) > 2:
            domain = ".".join(domain_parts[1:])
        else:
            domain = parsed_url.netloc

        # Regex for finding subdomains
        sub_domains = {
           
            'Sub-Domain': f'[^0-9_\-][0-9A-Za-z-_]+\.{domain}',
            'Sub-Domain': f'https?://[a-zA-Z0-9\.]+\.{domain}'
        }             

        if tool_method == "Path_and_Endpoints":

            print(f'\n{MAGENTA}[+] Testing {url}{RESET}')

            endpoint_check += "good"
            host = match(api_extractor, content, url, unique_host, sheet, wb, matched_patterns, tool_method, static_files, args, final_xlsx)
            if host is not None:
                final_host = host

        if tool_method == "URLs":
            print(f'\n{MAGENTA}[+] Testing URL: {url}{RESET}')
            host = match(uri_finder, content, url, unique_host, sheet, wb, matched_patterns, tool_method, static_files, args, final_xlsx)
            
            if host is not None:
                final_host = host

        if tool_method == "Sub-Domains":

            print(f'\n{MAGENTA}[+] Testing {url}{RESET}')
            host = match(sub_domains, content, url, unique_host, sheet, wb, matched_patterns, tool_method, static_files, args, final_xlsx)
            if host is not None:
                final_host = host

        if tool_method == "Secrets":
            print(f'\n{MAGENTA}[+] Testing {url}{RESET}')
            host = match(regex_secrets, content, url, unique_host, sheet, wb, matched_patterns, tool_method, static_files, args, final_xlsx)
            if host is not None:
                final_host = host
    
    if final_host is not None:
        filename = f'{final_host}\{final_host}_{tool_method}.xlsx'

        if os.path.isfile(filename):
            print(f'\n{GREEN}[+] {final_host}_{tool_method}.xlsx was created in your current directory!{RESET}')
        
        elif endpoint_check == "":
            print(f'\n{RED}[-] Nothing found for {tool_method} in {final_host}{RESET}')
        else:
            print(f'\n{RED}[-] Nothing found for {tool_method} in {final_host}{RESET}')

    elif final_host is None:
        if endpoint_check == "":
            print(f'\n{RED}[-] Nothing found for {tool_method} in {unique_host}{RESET}')
        else:
            print(f'\n{RED}[-] Nothing found for {tool_method} in {unique_host}{RESET}')


if __name__ == '__main__':

    start = time.time()

    args = parse_args()

    # Create a multiprocessing manager
    manager = multiprocessing.Manager()

    # Create a shared list using the manager
    final_xlsx = manager.list()

    # Get the number of CPU cores
    num_processes = args.threads

    # Create a pool of processes
    pool = multiprocessing.Pool(processes=num_processes)

    static_files = [
    
    "/css",
    "/assets",
    "/images",
    "/en-us",
    "/js",
    "/fonts",
    "/videos",
    "/audio",
    "/icons",
    "/favicon.ico",
    "/robots.txt",
    "/styles",
    "/media",
    "/docs",
    "/thumbnails",
    "/gallery",
    "/logos",
    "/banners",
    "/carousel",
    "/header",
    "/footer",
    "/slider",
    "/buttons",
    "/navigation",
    "/menu",
    "/placeholders",
    "/backgrounds",
    "/headers",
    "/footers",
    "/sidebars",
    "/widgets",
    "/social-media",
    "/error-pages",
    "/404",
    "/403",
    "/500",
    "/maintenance",
    "/analytics",
    "/tracking",
    "/ajax",
    "/rss",
    "/sitemap",
    "/feeds",
    "/webmanifest",
    "/config",
    "/locales",
    "/lib",
    "/vendor",
    "/build",
    "/dist",
    "/src",
    "/bower_components",
    "/node_modules",
    "/images",
    "/videos",
    "/documents",
    "/archives",
    "/fonts",
    "/stylesheets",
    "/plugins",
    "/themes",
    "/templates",
    "/includes",
    "/layouts",
    "/partials",
    "/locales",
    "/routes",
    "/controllers",
    "/models",
    "/views",
    "/helpers",
    "/middleware",
    'af-ZA', 
    'ar',     # Arabic
    'bg-BG',  # Bulgarian (Bulgaria)
    'ca-ES',  # Catalan (Spain)
    'cs-CZ',  # Czech (Czech Republic)
    'da-DK',  # Danish (Denmark)
    'de-DE',  # German (Germany)
    'el-GR',  # Greek (Greece)
    'en-US',  # English (United States)
    'es-ES',  # Spanish (Spain)
    'et-EE',  # Estonian (Estonia)
    'fi-FI',  # Finnish (Finland)
    'fr-FR',  # French (France)
    'he-IL',  # Hebrew (Israel)
    'hi-IN',  # Hindi (India)
    'hr-HR',  # Croatian (Croatia)
    'hu-HU',  # Hungarian (Hungary)
    'id-ID',  # Indonesian (Indonesia)
    'it-IT',  # Italian (Italy)
    'ja-JP',  # Japanese (Japan)
    'ko-KR',  # Korean (South Korea)
    'lt-LT',  # Lithuanian (Lithuania)
    'lv-LV',  # Latvian (Latvia)
    'ms-MY',  # Malay (Malaysia)
    'nb-NO',  # Norwegian Bokmål (Norway)
    'nl-NL',  # Dutch (Netherlands)
    'pl-PL',  # Polish (Poland)
    'pt-BR',  # Portuguese (Brazil)
    'pt-PT',  # Portuguese (Portugal)
    'ro-RO',  # Romanian (Romania)
    'ru-RU',  # Russian (Russia)
    'sk-SK',  # Slovak (Slovakia)
    'sl-SI',  # Slovenian (Slovenia)
    'sr-RS',  # Serbian (Serbia)
    'sv-SE',  # Swedish (Sweden)
    'th-TH',  # Thai (Thailand)
    'tr-TR',  # Turkish (Turkey)
    'uk-UA',  # Ukrainian (Ukraine)
    'vi-VN',  # Vietnamese (Vietnam)
    'zh-CN',  # Chinese (Simplified, China)
    'zh-TW',  # Chinese (Traditional, Taiwan)
]
    
   # Some regex for finding intersting stuff
    regex_secrets = {
    'google_api'     : r'AIza[0-9A-Za-z-_]{35}',
    'firebase'  : r'AAAA[A-Za-z0-9_-]{7}:[A-Za-z0-9_-]{140}',
    'google_captcha' : r'6L[0-9A-Za-z-_]{38}|^6[0-9a-zA-Z_-]{39}$',
    'google_oauth'   : r'ya29\.[0-9A-Za-z\-_]+',
    'amazon_aws_access_key_id' : r'A[SK]IA[0-9A-Z]{16}',
    'amazon_mws_auth_toke' : r'amzn\\.mws\\.[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}',
    'amazon_aws_url' : r's3\.amazonaws.com[/]+|[a-zA-Z0-9_-]*\.s3\.amazonaws.com',
    'amazon_aws_url2' : r"(" \
            r"[a-zA-Z0-9-\.\_]+\.s3\.amazonaws\.com" \
            r"|s3://[a-zA-Z0-9-\.\_]+" \
            r"|s3-[a-zA-Z0-9-\.\_\/]+" \
            r"|s3.amazonaws.com/[a-zA-Z0-9-\.\_]+" \
            r"|s3.console.aws.amazon.com/s3/buckets/[a-zA-Z0-9-\.\_]+)",
    'facebook_access_token' : r'EAACEdEose0cBA[0-9A-Za-z]+',
    'authorization_basic' : r'basic [a-zA-Z0-9=:_\+\/-]{5,100}',
    'authorization_bearer' : r'bearer [a-zA-Z0-9_\-\.=:_\+\/]{5,100}',
    'authorization_api' : r'api[key|_key|\s+]+[a-zA-Z0-9_\-]{5,100}',
    'paypal_braintree_access_token' : r'access_token\$production\$[0-9a-z]{16}\$[0-9a-f]{32}',
    'square_oauth_secret' : r'sq0csp-[ 0-9A-Za-z\-_]{43}|sq0[a-z]{3}-[0-9A-Za-z\-_]{22,43}',
    'square_access_token' : r'sqOatp-[0-9A-Za-z\-_]{22}|EAAA[a-zA-Z0-9]{60}',
    'stripe_standard_api' : r'sk_live_[0-9a-zA-Z]{24}',
    'stripe_restricted_api' : r'rk_live_[0-9a-zA-Z]{24}',
    'github_access_token' : r'[a-zA-Z0-9_-]*:[a-zA-Z0-9_\-]+@github\.com*',
    'rsa_private_key' : r'-----BEGIN RSA PRIVATE KEY-----',
    'ssh_dsa_private_key' : r'-----BEGIN DSA PRIVATE KEY-----',
    'ssh_dc_private_key' : r'-----BEGIN EC PRIVATE KEY-----',
    'pgp_private_block' : r'-----BEGIN PGP PRIVATE KEY BLOCK-----',
    'json_web_token' : r'ey[A-Za-z0-9-_=]+\.[A-Za-z0-9-_=]+\.?[A-Za-z0-9-_.+/=]*$',
    'slack_token' : r"\"api_token\":\"(xox[a-zA-Z]-[a-zA-Z0-9-]+)\"",
    'SSH_privKey' : r"([-]+BEGIN [^\s]+ PRIVATE KEY[-]+[\s]*[^-]*[-]+END [^\s]+ PRIVATE KEY[-]+)",
    }

    # Regex for finding Paths and APIs Endpoints *
    api_extractor = { 
    'PATH Finder-v1': r'\/[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder-v2': r'[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder-v3': r'\/api[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder-v4': r'api[a-zA-Z_]*(?:\/[a-zA-Z-_]+)*\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder-v5': r'\/[a-zA-Z]+(?:\/[a-zA-Z]+)+\/(?=[a-zA-Z]+\/)?[a-zA-Z]+(?:\/[a-zA-Z]+)*',
    'PATH Finder-v6': r'\/[a-zA-Z0-9]+(?:\/[a-zA-Z-0-9]+)*(?:\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
    'PATH Finder-v7': r'[a-zA-Z0-9]+(?:\/[a-zA-Z-0-9]+)*(?:\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
    'PATH Finder Backslash-1': r'\/[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder Backslash-v2': r'[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder Backslash-v3': r'\/api[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder Backslash-v4': r'api[a-zA-Z_]*(?:\\\/[a-zA-Z-_]+)*\\\/[a-zA-Z0-9_]*(?:-[a-zA-Z0-9_]+)*(?:-[a-zA-Z0-9_]+)*[a-zA-Z0-9_]*[a-zA-Z0-9_]*[a-zA-Z0-9_]',
    'PATH Finder Backslash-v5': r'\\\/[a-zA-Z]+(?:\\\/[a-zA-Z]+)+\\\/(?=[a-zA-Z]+\\\/)?[a-zA-Z]+(?:\\\/[a-zA-Z]+)*',
    'PATH Finder Backslash-v6': r'\\\/[a-zA-Z0-9]+(?:\\\/[a-zA-Z-0-9]+)*(?:\\\\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
    'PATH Finder Backslash-v7': r'[a-zA-Z]+(?:\\\/[a-zA-Z]+)+\\\/(?=[a-zA-Z]+\\\/)?[a-zA-Z]+(?:\\\/[a-zA-Z]+)*',
    'PATH Finder Backslash-v8': r'[a-zA-Z0-9]+(?:\\\/[a-zA-Z-0-9]+)*(?:\\\\/[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*\.[a-zA-Z0-9]+)',
    }

    # Regex for finding URLs
    uri_finder = {
        'URL Finder': '(http|ftp|https):\/\/([\w_-]+(?:(?:\.[\w_-]+)+))([\w.,@?^=%&:\/~+#-]*[\w@?^=%&\/~+#-])',
        'URL Finder Backslash': '(http|ftp|https):\\\/\\\/([\w_-]+(?:(?:\.[\w_-]+)+))([\w.,@?^=%&:\\\/~+#-]*[\w@?^=%&\\\/~+#-])'
    }
    
    warnings.filterwarnings("ignore")

    files = []    
    matched_patterns = set()
    path_js = set()
    task_args = []

    # Creating a new Workbook object
    wb = Workbook()
    wb_json = Workbook()

    if args.directory and args.file:
        print(f'\n{RED}[-] Choose either --file or --directory not both!{RESET}')
        exit(1)

    elif args.file:
        filename = args.file

    elif args.directory:
        directory = os.fsencode(args.directory)            
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            files.append(os.path.join(directory.decode(), filename))
            if "." in filename or filename.endswith(".py") or filename.endswith(".txt"): 
                continue
     
    if args.all and args.secrets or args.all and args.api or args.all and args.urls or args.all and args.bitrix or args.all and args.json and args.js and args.wordlist:
        print(f'\n{RED}If --all is set, remove other arguments(api/secrets/urls/bitrix/json/postman/js/domain).{RESET}')
        exit(1)

    if args.all:

        if args.directory:
           
           for filename in files:

                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break
                
                sheet_secrets, wb_secrets = create_worksheet(host, "Secrets")
                
                sheet_api_finder, wb_api_finder = create_worksheet(host,"Path_and_Endpoints")

                sheet_url_finder, wb_url_finder = create_worksheet(host, "URLs")

                sheet_sub_domains, wb_sub_domains = create_worksheet(host, "Sub-Domains")

                sheet_js, wb_js = create_worksheet_js("JS-Files")

                if not args.verbose:
                        print(f'{BLUE}\n[+] Executing Bitrix, Postman, JSON files, Secrets, URLs, APIs in JS files and Sub-domains methods for {host}...{RESET}')  
                        print(f'{BLUE}[+] Add --verbose to see the output printed to the screen with colors.{RESET}')
                        print(f'\n{BLUE}[+] This might take a while, be patient I tell you!{RESET}')
                    
                print(f'{BLUE}\n[+] Testing Bitrix method for {host}...{RESET}')  
                bitrix(filename)
                print(f'{BLUE}\n[+] Testing Postman method for {host}...{RESET}')  
                postMan(filename)
                print(f'{BLUE}\n[+] Testing JSON method for {host}...{RESET}') 
                json_file(filename, wb_json)
                print(f'{BLUE}\n[+] Testing JS URLs method for {host}...{RESET}') 
                js_file(filename, wb_js, sheet_js)
                print(f'{BLUE}\n[+] Testing URLs method for {host}...{RESET}') 
                task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
                print(f'{BLUE}\n[+] Testing Secrets method for {host}...{RESET}')  
                task_args.append((filename, "Secrets", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
                print(f'{BLUE}\n[+] Testing APIs & Paths method for {host}...{RESET}') 
                task_args.append((filename, "Path_and_Endpoints", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
                print(f'{BLUE}\n[+] Testing Sub-Domains method for {host}...{RESET}')  
                task_args.append((filename, "Sub-Domains", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
                print(f'{BLUE}\n[+] Creating wordlist tailored to {host}...{RESET}')  
                wordlist_creator(filename, host)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break
            
            if not args.verbose:
                    print(f'{BLUE}\n[+] Executing Bitrix, Postman, JSON files, Secrets, URLs, APIs in JS files and Sub-domains methods for {host}...{RESET}')  
                    print(f'{BLUE}[+] Add --verbose to see the output printed to the screen with colors.{RESET}')
                    print(f'\n{BLUE}[+] This might take a while, be patient I tell you!{RESET}')
            
            sheet_url_finder, wb_url_finder = create_worksheet_main("URLs")

            sheet_secrets, wb_secrets = create_worksheet_main("Secrets")

            sheet_api_finder, wb_api_finder = create_worksheet_main("API Endpoints")

            sheet_sub_domains, wb_sub_domains = create_worksheet_main("Sub Domains")
            
            sheet_js, wb_js = create_worksheet_js("JS-Files")
            
            print(f'{BLUE}\n[+] Testing Bitrix method for {host}...{RESET}')  
            bitrix(filename)
            print(f'{BLUE}\n[+] Testing Postman method for {host}...{RESET}')  
            postMan(filename)
            print(f'{BLUE}\n[+] Testing JSON method for {host}...{RESET}') 
            json_file(filename, wb_json)
            print(f'{BLUE}\n[+] Testing JS URLs method for {host}...{RESET}') 
            js_file(filename, wb_js, sheet_js)
            print(f'{BLUE}\n[+] Testing URLs method for {host}...{RESET}') 
            task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
            print(f'{BLUE}\n[+] Testing Secrets method for {host}...{RESET}')  
            task_args.append((filename, "Secrets", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
            print(f'{BLUE}\n[+] Testing APIs & Paths method for {host}...{RESET}') 
            task_args.append((filename, "Path_and_Endpoints", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
            print(f'{BLUE}\n[+] Testing Sub-Domains method for {host}...{RESET}')  
            task_args.append((filename, "Sub-Domains", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
            print(f'{BLUE}\n[+] Creating wordlist tailored to {host}...{RESET}')  
            wordlist_creator(filename, host)
            
    if not args.all:

        sheet_url_finder, wb_url_finder = create_worksheet_main("URLs")

        sheet_secrets, wb_secrets = create_worksheet_main("Secrets")

        sheet_api_finder, wb_api_finder = create_worksheet_main("API Endpoints")

        sheet_sub_domains, wb_sub_domains = create_worksheet_main("Sub Domains")

        sheet_js, wb_js = create_worksheet_js("JS-Files")

    if args.bitrix and not args.all:
        
        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break
                
                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing Bitrix method for {host}...{RESET}')  
                
                bitrix(filename)

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:  
                print(f'{BLUE}\n[+] Testing Bitrix method for {host}...{RESET}')  

            bitrix(filename)

    if args.postman and not args.all:
        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing Postman method for {host}...{RESET}')  
                
                postMan(filename)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:      
                print(f'{BLUE}\n[+] Testing Postman method for {host}...{RESET}')  
            
            postMan(filename)

    if args.js and not args.all:
        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing JS URLs method for {host}...{RESET}')   
                
                js_file(filename, wb_js, sheet_js)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:      
                print(f'{BLUE}\n[+] Testing JS URLs method for {host}...{RESET}')  
            
            js_file(filename, wb_js, sheet_js)
            
    if args.urls and not args.all:
        
        counter = 0
        if args.directory:
            for filename in files:
                counter += 1
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break
                
                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing URLs method for {host}...{RESET}')  

                sheet_url_finder, wb_url_finder = create_worksheet_main(f"URLs_{counter}")

                task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

                
        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:      
                print(f'{BLUE}\n[+] Testing URLs method for {host}...{RESET}')  

            task_args.append((filename, "URLs", sheet_url_finder, wb_url_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

    if args.domain and not args.all:

        if args.directory:
            counter = 0
            for filename in files:
                counter += 1
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break
                
                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing Sub-Domains method for {host}...{RESET}')  
                
                sheet_sub_domains, wb_sub_domains = create_worksheet_main(f"Sub Domains_{counter}")
                # Map the function and arguments to the pool
                task_args.append((filename, "Sub-Domains", sheet_sub_domains, wb_sub_domains, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))


        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:     
                print(f'{BLUE}\n[+] Testing Sub-Domains method for {host}...{RESET}')  

            # Map the function and arguments to the pool
            task_args.append((filename, "Sub-Domains", sheet_sub_domains, wb_sub_domains, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

    if args.json and not args.all:
        # Creating a new Workbook object
        wb_json = Workbook()

        if args.directory:
            for filename in files:
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break
                
                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing JSON method for {host}...{RESET}')  
                
                json_file(filename, wb_json)

        elif not args.directory:
            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:      
                print(f'{BLUE}\n[+] Testing JSON method for {host}...{RESET}')  

            json_file(filename, wb_json)

    if args.api and not args.all:

        if args.directory:
            
            counter = 0
            for filename in files:
                counter += 1
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break

                if not args.all and not args.verbose:    
                    print(f'\n{BLUE}[+] This might take a while, be patient I tell you!{RESET}')    
                    print(f'{BLUE}\n[+] Testing APIs and PATHs with REGEX method for {host}...{RESET}')

                sheet_api_finder, wb_api_finder = create_worksheet_main(f"API Endpoints_{counter}")

                # Map the function and arguments to the pool
                task_args.append((filename, "Path_and_Endpoints", sheet_api_finder, wb_api_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break
            
            if not args.all and not args.verbose:  
                print(f'{BLUE}\n[+] Testing APIs and PATHs with REGEX method for {host}...{RESET}')  
                print(f'\n{BLUE}[+] This might take a while, be patient I tell you!{RESET}')
            
            # Map the function and arguments to the pool
            task_args.append((filename, "Path_and_Endpoints", sheet_api_finder, wb_api_finder, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

    if args.secrets and not args.all:

        counter = 0
        argumnets = [args.urls, args.api, args.domain]
        if args.directory:
           for filename in files:   
                # Create an XML Object
                tree = ET.parse(filename)
                main_root = tree.getroot()

                for i in main_root:
                    response = i.find('response').text
                    if response is None:
                        continue
                    content = base64.b64decode(response)
                    content = content.decode('latin-1')
                    host = i.find('host').text
                    break
                
                if not args.all and not args.verbose:  
                    print(f'{BLUE}\n[+] Testing Secrets method with REGEX for {host}...{RESET}')  

                sheet_secrets, wb_secrets = create_worksheet_main(f"Secrets_{counter}")
                # Map the function and arguments to the pool
                task_args.append((filename, "Secrets", sheet_secrets, wb_secrets, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))
       
        elif not args.directory:

            # Create an XML Object
            tree = ET.parse(filename)
            main_root = tree.getroot()

            for i in main_root:
                response = i.find('response').text
                if response is None:
                    continue
                content = base64.b64decode(response)
                content = content.decode('latin-1')
                host = i.find('host').text
                break

            if not args.all and not args.verbose:  
                print(f'{BLUE}\n[+] Testing Secrets method with REGEX for {host}...{RESET}') 

            task_args.append((filename, "Secrets", sheet_secrets, wb_secrets, uri_finder, static_files, regex_secrets, api_extractor, args, matched_patterns, final_xlsx))

    if args.wordlist and not args.all:

        # Create an XML Object
        tree = ET.parse(filename)
        main_root = tree.getroot()

        for i in main_root:
            response = i.find('response').text
            if response is None:
                continue
            content = base64.b64decode(response)
            content = content.decode('latin-1')
            host = i.find('host').text
            break
        
        if not args.all and not args.verbose:  
            print(f'{BLUE}\n[+] Creating wordlist tailored to {host}...{RESET}')  

        wordlist_creator(filename, host)

    # Create a pool of processes using a context manager
    with multiprocessing.Pool(processes=num_processes) as pool:
        # Start the processes
        results = pool.starmap_async(main, task_args)    
        
        # Wait for the results 
        while not results.ready():
            time.sleep(0)

    # Terminate all processes
    pool.terminate()
    pool.join()

    check_dup = []

    for excel in final_xlsx:
        if excel not in check_dup:
            check_dup.append(excel)
            adjust_column_widths_disk(excel)

        if excel in check_dup:
            continue
    
    end = time.time()
    hours, rem = divmod(end-start, 3600)
    minutes, seconds = divmod(rem, 60)        
    print("\n--- Running time: {:0>2}:{:0>2}:{:05.2f}".format(int(hours),int(minutes),seconds), "---")
